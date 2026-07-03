from openpyxl import load_workbook
from fastapi import UploadFile, HTTPException
from sqlmodel import Session, select

from app.models.topic import Topic
from app.models.question import Question
from app.services.question_service import bulk_save_questions


VALID_LEVELS = {
    "beginner",
    "intermediate",
    "advanced"
}

VALID_ANSWERS = {
    "a",
    "b",
    "c",
    "d"
}


def upload_questions_from_excel(
    db: Session,
    file: UploadFile,
    created_by: int
):

    # ==========================================
    # File Validation
    # ==========================================

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx files are allowed."
        )

    workbook = load_workbook(file.file)

    sheet = workbook.active

    expected_headers = [

        "Topic",
        "Level",
        "Question",
        "Option A",
        "Option B",
        "Option C",
        "Option D",
        "Correct Answer",
        "Explanation"

    ]

    headers = [
        cell.value
        for cell in sheet[1]
    ]

    if headers != expected_headers:

        raise HTTPException(

            status_code=400,

            detail="Invalid Excel format. Download the template first."

        )

    questions = []

    errors = []

    duplicate_excel = set()

    # ==========================================
    # Read Rows
    # ==========================================

    for row_number, row in enumerate(

        sheet.iter_rows(
            min_row=2,
            values_only=True
        ),

        start=2

    ):

        (

            topic_name,
            level,
            question,
            option_a,
            option_b,
            option_c,
            option_d,
            correct_answer,
            explanation

        ) = row

        # ==========================================
        # Empty Question
        # ==========================================

        if not question:

            errors.append({

                "row": row_number,

                "reason": "Question is empty"

            })

            continue

        # ==========================================
        # Duplicate inside Excel
        # ==========================================

        duplicate_key = (

            str(topic_name).lower(),

            str(question).strip().lower()

        )

        if duplicate_key in duplicate_excel:

            errors.append({

                "row": row_number,

                "reason": "Duplicate question inside Excel"

            })

            continue

        duplicate_excel.add(duplicate_key)

        # ==========================================
        # Topic Validation
        # ==========================================

        topic = db.exec(

            select(Topic).where(

                Topic.name.ilike(
                    str(topic_name)
                )

            )

        ).first()

        if not topic:

            errors.append({

                "row": row_number,

                "reason": f"Topic '{topic_name}' not found"

            })

            continue

        # ==========================================
        # Level Validation
        # ==========================================

        level = str(level).lower().strip()

        if level not in VALID_LEVELS:

            errors.append({

                "row": row_number,

                "reason": "Invalid level"

            })

            continue

        # ==========================================
        # Correct Answer Validation
        # ==========================================

        correct_answer = str(

            correct_answer

        ).lower().strip()

        if correct_answer not in VALID_ANSWERS:

            errors.append({

                "row": row_number,

                "reason": "Correct answer must be A/B/C/D"

            })

            continue

        # ==========================================
        # Empty Options
        # ==========================================

        if not all([

            option_a,

            option_b,

            option_c,

            option_d

        ]):

            errors.append({

                "row": row_number,

                "reason": "All options are required"

            })

            continue

        # ==========================================
        # Duplicate Database
        # ==========================================

        existing = db.exec(

            select(Question).where(

                Question.topic_id == topic.id,

                Question.question == question

            )

        ).first()

        if existing:

            errors.append({

                "row": row_number,

                "reason": "Question already exists in database"

            })

            continue

        questions.append({

            "topic_id": topic.id,

            "level": level,

            "question": question.strip(),

            "option_a": option_a.strip(),

            "option_b": option_b.strip(),

            "option_c": option_c.strip(),

            "option_d": option_d.strip(),

            "correct_answer": correct_answer,

            "explanation": explanation,

            "source": "excel",

            "created_by": created_by

        })

    # ==========================================
    # Bulk Save
    # ==========================================

    result = bulk_save_questions(

        db,

        questions

    )

    return {

        "total_rows": sheet.max_row - 1,

        "inserted": result["inserted"],

        "duplicates": result["duplicates"],

        "failed": len(errors),

        "errors": errors

    }