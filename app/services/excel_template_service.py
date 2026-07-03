from openpyxl import Workbook
from io import BytesIO


def generate_question_template():

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Questions"

    headers = [

        "Topic",
        "Level",
        "Question",
        "Option A",
        "Option B",
        "Option C",
        "Option D",
        "Correct Answer",
        "Explanation"

    ]

    for column, header in enumerate(headers, start=1):

        sheet.cell(
            row=1,
            column=column
        ).value = header

    # ==========================================
    # Sample Row
    # ==========================================

    sample = [

        "Java",

        "beginner",

        "JVM stands for?",

        "Java Virtual Machine",

        "Java Variable Machine",

        "Java Version Machine",

        "None",

        "A",

        "JVM executes Java bytecode."

    ]

    for column, value in enumerate(sample, start=1):

        sheet.cell(
            row=2,
            column=column
        ).value = value

    stream = BytesIO()

    workbook.save(stream)

    stream.seek(0)

    return stream